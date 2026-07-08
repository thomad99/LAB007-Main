#!/usr/bin/env python3
"""Import Elite-Client-List.xlsx (Client List NEW sheet) to JSON for Elite Invoices ClientFull tab."""

from __future__ import annotations

import json
import sys
from datetime import date, datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = Path.home() / "OneDrive" / "Elite Cleaning" / "Elite-Client-List.xlsx"
OUT_PATH = ROOT / "data" / "elite-client-list-new.json"
SHEET_NAME = "Client List NEW"


def normalize(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value == int(value):
        return int(value)
    return value


def main() -> int:
    try:
        import openpyxl
    except ImportError:
        print("Install openpyxl: pip install openpyxl", file=sys.stderr)
        return 1

    xlsx_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    if not xlsx_path.exists():
        print(f"File not found: {xlsx_path}", file=sys.stderr)
        return 1

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        print(f"Sheet not found: {SHEET_NAME}", file=sys.stderr)
        return 1

    ws = wb[SHEET_NAME]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    records = []
    for row in rows[1:]:
        if not any(cell not in (None, "") for cell in row):
            continue
        item = {}
        for index, header in enumerate(headers):
            if not header:
                continue
            item[header] = normalize(row[index] if index < len(row) else "")
        if not str(item.get("Name", "")).strip() and not str(item.get("Client ID", "")).strip():
            continue
        records.append(item)

    payload = {"sheet": SHEET_NAME, "columns": headers, "rows": records}
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUT_PATH.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"Wrote {len(records)} rows to {OUT_PATH}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
